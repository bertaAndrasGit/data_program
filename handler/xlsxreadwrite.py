import sqlite3
from openpyxl.workbook import Workbook

from models.item import Item
from models.worker import Worker
from models.storage import Storage

class Xlsx_rw:
    @staticmethod
    def write(tablename:str,wb: Workbook,heading: bool = True) -> None:
        con = sqlite3.connect(r'..\..\beadando\db\storage.db')
        cur = con.cursor()

        cur.execute(f"SELECT * FROM {tablename}")
        columns = [desc[0] for desc in cur.description]

        sheet = wb.create_sheet(tablename)
        if heading:
            for col in range(len(columns)):
                sheet.cell(row=1,column=col+1,value=columns[col])

        columns_count = len(columns)
        offset = 2 if heading else 1
        res = cur.execute(f"SELECT count(*) FROM {tablename}")
        length = res.fetchone()
        cur.execute(f"SELECT * FROM {tablename}")

        data = cur.fetchall()

        #print(length[0])
        for row in range(length[0]):
            for column in range(1,columns_count+1):
                sheet.cell(row=row + offset,column=column ,value=data[row][column-1])


        cur.close()
        con.close()

    @staticmethod
    def read(entity,wb: Workbook, sheet_name: str = "Workers",heading:bool = True) -> list[object]:
        con = sqlite3.connect(r'..\..\beadando\db\storage.db')
        cur = con.cursor()
        sheet = wb[sheet_name]

        objects = []
        row = 2 if heading else 1
        while True:
            cell = sheet.cell(row=row,column=1)
            if cell.value is None:
                break

            cur.execute(f"SELECT * FROM {sheet_name.lower()}")
            columns = [desc[0] for desc in cur.description]
            #print(columns)
            columns_count = len(columns)

            sheet_row = [sheet.cell(row=row,column=column).value for column in range(1,columns_count+1)]
            #print(*sheet_row)
            objects.append(entity(*sheet_row))
            row += 1



        cur.close()
        con.close()
        return objects


if __name__ == '__main__':
    wb = Workbook()

    Xlsx_rw.write("Workers", wb)
    print(Xlsx_rw.read(Worker,wb,"Workers"))
    print("-----------------------")
    Xlsx_rw.write("Items", wb)
    print(Xlsx_rw.read(Item,wb,"Items"))
    print("-----------------------")
    Xlsx_rw.write("Storages", wb)
    print(Xlsx_rw.read(Storage,wb,"Storages"))

    wb.save("./database.xlsx")